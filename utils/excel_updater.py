import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
import os


def update_excel_status(
        test_case_id: str,
        status: str,
        duration: float,
        error_message: Optional[str] = None,
        screenshot_path: Optional[str] = None,
        file_path: str = 'Test_Cases.xlsx',
        sheet_name: Optional[str] = None
):
    """
    Finds a test case by its ID in an Excel sheet and updates its status.
    This version cleans up the error message and adds a screenshot link.
    """
    full_path = Path.cwd() / file_path
    if not full_path.exists():
        print(f"Error: Excel file not found at {full_path}")
        return

    workbook = openpyxl.load_workbook(full_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # --- Find Column Headers ---
    header = [cell.value for cell in sheet[1]]
    try:
        id_col_idx = header.index("Test Case ID") + 1
        status_col_idx = header.index("Execution Status") + 1
    except ValueError as e:
        print(f"Error: Column not found in Excel sheet '{sheet.title}': {e}")
        return

    # --- LOGIC TO CLEAN UP THE ERROR MESSAGE ---
    actual_result_message = 'Test Passed Successfully'
    if status.lower() == 'failed' and error_message:
        lines = error_message.strip().split('\n')
        error_lines = [line.strip().replace('E   ', '') for line in lines if line.strip().startswith('E   ')]
        if error_lines:
            actual_result_message = ' | '.join(error_lines)
        else:
            actual_result_message = lines[-1].strip()
    elif status.lower() == 'skipped' and error_message:
        actual_result_message = f"Skipped: {error_message}"

    # --- Add new columns for results if they don't exist ---
    new_headers = {
        "Timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "Duration (s)": f'{duration:.4f}',
        "Actual Result": actual_result_message,
        "Screenshot": "N/A"
    }

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")

    for col_name in new_headers.keys():
        if col_name not in header:
            new_col_idx = sheet.max_column + 1
            cell = sheet.cell(row=1, column=new_col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            # Made the "Actual Result" column wider for better readability
            if col_name == "Actual Result":
                sheet.column_dimensions[cell.column_letter].width = 80
            elif col_name == "Screenshot":
                sheet.column_dimensions[cell.column_letter].width = 25
            else:
                sheet.column_dimensions[cell.column_letter].width = 25

    # --- Find the Row and Update It ---
    target_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=id_col_idx).value == test_case_id:
            target_row = row
            break

    if not target_row:
        print(f"Warning: Test Case ID '{test_case_id}' not found in sheet '{sheet.title}'.")
        return

    # Update the status cell
    status_cell = sheet.cell(row=target_row, column=status_col_idx)
    status_cell.value = status.capitalize()

    # Apply color coding
    if status.lower() == 'passed':
        status_cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        status_cell.font = Font(color="FF006100")
    elif status.lower() == 'failed':
        status_cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        status_cell.font = Font(color="FF9C0006")

    # Update the new result columns
    current_header = [cell.value for cell in sheet[1]]  # Re-read header
    for col_name, value in new_headers.items():
        try:
            col_idx = current_header.index(col_name) + 1
            if col_name == "Screenshot":
                # Write the default "N/A"
                if not screenshot_path:
                    sheet.cell(row=target_row, column=col_idx, value=value)
            else:
                sheet.cell(row=target_row, column=col_idx, value=value)
        except ValueError:
            continue

    if screenshot_path:
        try:
            # Find the "Screenshot" column
            screenshot_col_idx = current_header.index("Screenshot") + 1
            screenshot_cell = sheet.cell(row=target_row, column=screenshot_col_idx)

            # Create a full, absolute file path and format it as a hyperlink
            # .replace handles Windows paths
            full_path = "file:///" + os.path.abspath(screenshot_path).replace("\\", "/")

            screenshot_cell.value = "View Screenshot"
            screenshot_cell.hyperlink = full_path
            screenshot_cell.style = "Hyperlink"
        except ValueError:
            print("Warning: 'Screenshot' column not found. Could not write hyperlink.")

    workbook.save(full_path)